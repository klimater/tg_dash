import openpyxl
import openpyxl.reader
import openpyxl.reader.excel

# Функция безопасного открытия файла
def open_exel(file_name):
    try:
        wb = openpyxl.reader.excel.load_workbook(filename= file_name)
        wb.active = 0
    except:
        print("Файл не найден")
        return None
    
    return wb

# Создание новой продажи
def add_sale(wb, date, name, category, quantity, cost):
    if wb is None:
        return "Ошибка файл не найден"
    
    sheet = wb.active

    for row in range(1, sheet.max_row + 2):
        if sheet[row][0].value == None:
            sheet[row][0].value = date
            sheet[row][1].value = name
            sheet[row][2].value = category
            sheet[row][3].value = quantity
            sheet[row][4].value = cost
            wb.save("BD.xlsx")
            return "Продажа зафиксирована"
        

# Удаление продажи
def delete_sale(wb, date, name, count_elem):
    if wb is None:
        return "Ошибка файл не найден"
    
    sheet = wb.active
    
    for row in range(1, sheet.max_row + 1):
        # поиск нужной продажи
        if str(sheet[row][0].value)[:10] == date and sheet[row][1].value == name:
            # проверка кол-ва удаляемых элементов
            if sheet[row][3].value != 0 and int(sheet[row][3].value) >= count_elem:
                sheet[row][3].value = int(sheet[row][3].value) - count_elem
                wb.save("BD.xlsx")
                return "Запись о продаже удалена"
            else:
                return "Ошибка: кол-во товаров превышает кол-ва проданых товаров"
        else:
            return "Товар не найден"


# Функция вывода даных из exel файла
def output_sheet(wb):
    if wb is None:
        return "Ошибка файл не найден"
    
    sheet = wb.active

    value = """"""
    for row in range(1, sheet.max_row + 2):
        date_sale = sheet[row][0].value
        name = sheet[row][1].value
        category = sheet[row][2].value
        quantity = sheet[row][3].value
        cost = sheet[row][4].value
        value += f"""{str(date_sale)[:11]} {name} {category} {quantity} {cost}\n"""
    return value

wb = open_exel("BD.xlsx")
delete_sale(wb, "03.12.2024", 'Ноутбук', 1)